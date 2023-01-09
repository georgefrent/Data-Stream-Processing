from datetime import datetime
from turtle import clear
import time
import requests
from openpyxl import load_workbook, Workbook
import pandas as pd

def p(titlu = "", *args):
    print('-'*80)
    print(titlu)
    print(*args, sep="\n")

print('Introduceti o valoare care va fi folosita la Cerinta 2:')
valoareConfigurabila = float(input())

currentMonth = datetime.now().month
currentYear = datetime.now().year
currentDay = datetime.now().day

switcher = {
        1: "ianuarie",
        2: "februarie",
        3: "martie",
        4: "aprilie",
        5: "mai",
        6: "iunie",
        7: "iulie",
        8: "august",
        9: "septembrie",
        10: "octombrie",
        11: "noiembrie",
        12: "decembrie"
    }

currentMonthString = switcher.get(currentMonth, "nothing")
currentYearString = str(currentYear)
currentDayString = str(currentDay)

theDayBefore = currentDay - 1

print(currentDay)
print (switcher.get(currentMonth, "nothing"))
print(currentYear)

if currentMonth < 10:
    currentMonth2 = '%02d' % currentMonth
else: currentMonth2 = currentMonth
currentMonth2String = str(currentMonth2)

if theDayBefore < 10:
    theDayBefore2 = '%02d' % theDayBefore
else: theDayBefore2 = theDayBefore
theDayBeforeString = str(theDayBefore2)


lastDay = currentYearString + '-' + currentMonth2String + '-' + theDayBeforeString
firstDay = currentYearString + '-' + currentMonth2String + '-01'


switcher = {
        "2022-"+currentMonth2String+"-01": 2,
        "2022-"+currentMonth2String+"-02": 3,
        "2022-"+currentMonth2String+"-03": 4,
        "2022-"+currentMonth2String+"-04": 5,
        "2022-"+currentMonth2String+"-05": 6,
        "2022-"+currentMonth2String+"-06": 7,
        "2022-"+currentMonth2String+"-07": 8,
        "2022-"+currentMonth2String+"-08": 9,
        "2022-"+currentMonth2String+"-09": 10,
        "2022-"+currentMonth2String+"-10": 11,
        "2022-"+currentMonth2String+"-11": 12,
        "2022-"+currentMonth2String+"-12": 13,
        "2022-"+currentMonth2String+"-13": 14,
        "2022-"+currentMonth2String+"-14": 15,
        "2022-"+currentMonth2String+"-15": 16,
        "2022-"+currentMonth2String+"-16": 17,
        "2022-"+currentMonth2String+"-17": 18,
        "2022-"+currentMonth2String+"-18": 19,
        "2022-"+currentMonth2String+"-19": 20,
        "2022-"+currentMonth2String+"-20": 21,
        "2022-"+currentMonth2String+"-21": 22,
        "2022-"+currentMonth2String+"-22": 23,
        "2022-"+currentMonth2String+"-23": 24,
        "2022-"+currentMonth2String+"-24": 25,
        "2022-"+currentMonth2String+"-25": 26,
        "2022-"+currentMonth2String+"-26": 27,
        "2022-"+currentMonth2String+"-27": 28,
        "2022-"+currentMonth2String+"-28": 29,
        "2022-"+currentMonth2String+"-29": 30,
        "2022-"+currentMonth2String+"-30": 31,
        "2022-"+currentMonth2String+"-31": 32,
    }

lastDayint = switcher.get(lastDay, "nothing")

#-----Cerinta 1----------------------------------

fileName = "transparenta_" + currentMonthString + "_" + currentYearString + ".xlsx"
p("Cerinta 1\nFisierul a fost descarcat cu succes sub numele: " + fileName)

urlMai2022 = 'https://data.gov.ro/dataset/b86a78a3-7f88-4b53-a94f-015082592466/resource/3bea5606-ef79-49bc-bdf7-c07e173429d6/download/transparenta_mai_2022.xlsx'
urlIunie2022 = 'https://data.gov.ro/dataset/b86a78a3-7f88-4b53-a94f-015082592466/resource/3bea5606-ef79-49bc-bdf7-c07e173429d6/download/transparenta_mai_2022.xlsx'

if currentMonth == 5:
    r = requests.get(urlMai2022, allow_redirects=True)
elif currentMonth == 6:
    r = requests.get(urlIunie2022, allow_redirects=True)

open(fileName, 'wb').write(r.content)

#-----End of cerinta 1--------------------------

#-----Cerinta 3----------------------------------

wb = load_workbook(fileName)
ws = wb['incidenta']
ws.delete_rows(1,2)
wb.save(fileName)

dataframe = pd.read_excel(fileName)

df = pd.DataFrame(dataframe, columns=['Judet','2022-05-22','2022-05-23','2022-05-24','2022-05-25','2022-05-26','2022-05-27','2022-05-28','2022-05-29','2022-05-30','2022-05-31',])

df2 = df.groupby('Judet').sum()

df2['Suma'] = df2['2022-05-22'] + df2['2022-05-23'] + df2['2022-05-24'] + df2['2022-05-25'] + df2['2022-05-26'] + df2['2022-05-27'] + df2['2022-05-28'] + df2['2022-05-29'] + df2['2022-05-30'] + df2['2022-05-31']

df3 = df2.sort_values(by=['Suma'], ascending=False)

df3Excel =  'Cerinta_3.xlsx'
df3.to_excel(df3Excel)

#-----Cerinta 2---------------------------------

#Am sters primele 3 randuri din fisierul excel pentru a putea accesa mai usor datele din celule

localitatiAlertate = list()

wb = load_workbook(fileName)
ws = wb['incidenta']
ws.delete_rows(1,1)
for row in ws.values:  
    crestere = row[lastDayint] - row[2]
    if crestere > valoareConfigurabila:
        localitatiAlertate.append(row[0])
wb.save(fileName)

p("Cerinta 2")
print(localitatiAlertate)
print('au cresterea incidentei mai mare decat ' + str(valoareConfigurabila) +" (Valoarea configurata)")

p("Raspunsul la Cerinta 3 se afla in fisierul cu numele Cerinta_3.xlsx")
